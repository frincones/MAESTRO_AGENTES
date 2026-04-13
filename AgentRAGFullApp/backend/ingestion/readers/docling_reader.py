"""Reader for document formats: PDF, DOCX, PPTX, XLSX, HTML.

Strategy: prefer FAST lightweight readers (pypdf, python-docx) for common cases.
Only use Docling for complex layouts (PPTX, XLSX, HTML with rich layout) since
Docling is heavy and slow on first call (loads ML models, several seconds cold start).
"""

from __future__ import annotations

import logging
import os
import time
from pathlib import Path
from typing import Optional, Tuple

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".xlsx", ".xls", ".html", ".htm"}
BINARY_EXTENSIONS = {".pdf", ".docx", ".doc", ".pptx", ".ppt", ".xlsx", ".xls"}

# Formats where the lightweight reader is sufficient and Docling adds no value worth its cost.
# Setting the env var FAST_INGESTION=0 disables this and forces Docling.
FAST_INGESTION = os.getenv("FAST_INGESTION", "1") != "0"
FAST_LIGHTWEIGHT_FORMATS = {".pdf", ".docx", ".xlsx", ".xls"}


def can_handle(file_path: str) -> bool:
    return Path(file_path).suffix.lower() in SUPPORTED_EXTENSIONS


def read(file_path: str) -> Tuple[str, Optional[object], dict]:
    """Read a document and return (markdown_text, docling_document_or_none, metadata)."""
    path = Path(file_path)
    ext = path.suffix.lower()
    metadata = {
        "reader": "docling_module",
        "format": ext,
        "file_name": path.name,
    }
    is_binary = ext in BINARY_EXTENSIONS
    t_start = time.time()

    # FAST PATH: for PDF and DOCX, try lightweight reader first
    if FAST_INGESTION and ext in FAST_LIGHTWEIGHT_FORMATS:
        result = _binary_fallback(file_path, dict(metadata), ext, raise_on_failure=False)
        if result is not None:
            elapsed = time.time() - t_start
            logger.info("Read %s with lightweight reader in %.2fs", path.name, elapsed)
            return result
        logger.info("Lightweight reader didn't work for %s, falling back to Docling", path.name)

    # SLOW PATH: Docling (handles layout, OCR, tables, etc.)
    try:
        from docling.document_converter import DocumentConverter

        converter = DocumentConverter()
        result = converter.convert(str(path))
        docling_doc = result.document
        markdown_content = result.document.export_to_markdown()

        if not markdown_content or not markdown_content.strip():
            logger.warning("Docling returned empty content for %s", path.name)
            if is_binary:
                return _binary_fallback(file_path, metadata, ext)
            return _fallback_read(file_path, metadata)

        metadata["pages"] = getattr(result.document, "num_pages", None)
        metadata["conversion"] = "docling"
        elapsed = time.time() - t_start
        logger.info("Read %s with Docling in %.2fs", path.name, elapsed)
        return markdown_content, docling_doc, metadata

    except ImportError:
        logger.info("Docling not installed. Using lightweight fallback for %s", path.name)
        if is_binary:
            return _binary_fallback(file_path, metadata, ext)
        return _fallback_read(file_path, metadata)
    except Exception as e:
        logger.warning("Docling failed for %s: %s. Using fallback.", path.name, e)
        if is_binary:
            return _binary_fallback(file_path, metadata, ext)
        return _fallback_read(file_path, metadata)


def _binary_fallback(
    file_path: str,
    metadata: dict,
    ext: str,
    raise_on_failure: bool = True,
):
    """Use lightweight format-specific readers (pypdf, python-docx).

    Returns (content, None, metadata) on success, or None if failed and
    raise_on_failure is False, or raises RuntimeError otherwise.
    """
    path = Path(file_path)

    if ext == ".pdf":
        try:
            import pypdf
            reader = pypdf.PdfReader(str(path))
            pages = []
            for i, page in enumerate(reader.pages, start=1):
                text = page.extract_text() or ""
                if text.strip():
                    pages.append(f"## Page {i}\n\n{text}")
            content = "\n\n".join(pages)
            if content.strip():
                metadata["conversion"] = "pypdf"
                metadata["pages"] = len(reader.pages)
                content = "\x00".join(content.split("\x00"))  # strip NULs
                content = content.replace("\x00", "")
                return f"# {path.name}\n\n{content}", None, metadata
        except ImportError:
            logger.warning("pypdf not installed for PDF fallback")
        except Exception as e:
            logger.warning("pypdf failed for %s: %s", path.name, e)

    if ext in (".docx",):
        try:
            from docx import Document as DocxDoc
            doc = DocxDoc(str(path))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            content = "\n\n".join(paragraphs)
            if content.strip():
                metadata["conversion"] = "python-docx"
                content = content.replace("\x00", "")
                return f"# {path.name}\n\n{content}", None, metadata
        except ImportError:
            logger.warning("python-docx not installed for DOCX fallback")
        except Exception as e:
            logger.warning("python-docx failed for %s: %s", path.name, e)

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
            sheets_md = []
            total_rows = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                total_rows += len(rows)

                # Header row + alignment
                header = [str(c) if c is not None else "" for c in rows[0]]
                if not any(header):
                    continue

                md = [f"## Sheet: {sheet_name}\n"]
                md.append("| " + " | ".join(header) + " |")
                md.append("| " + " | ".join(["---"] * len(header)) + " |")
                for row in rows[1:]:
                    cells = [
                        str(c).replace("\n", " ").replace("|", "\\|") if c is not None else ""
                        for c in row
                    ]
                    # Pad / trim row to header length
                    while len(cells) < len(header):
                        cells.append("")
                    cells = cells[: len(header)]
                    md.append("| " + " | ".join(cells) + " |")
                sheets_md.append("\n".join(md))
            wb.close()

            content = "\n\n".join(sheets_md)
            if content.strip():
                metadata["conversion"] = "openpyxl"
                metadata["sheets"] = len(sheets_md)
                metadata["total_rows"] = total_rows
                content = content.replace("\x00", "")
                return f"# {path.name}\n\n{content}", None, metadata
        except ImportError:
            logger.warning("openpyxl not installed for XLSX fallback")
        except Exception as e:
            logger.warning("openpyxl failed for %s: %s", path.name, e)

    if not raise_on_failure:
        return None
    raise RuntimeError(
        f"Cannot read {path.name}: no suitable reader available. "
        f"Install Docling (pip install docling) for full support, "
        f"or convert {ext} files to .md/.txt before uploading."
    )


def _fallback_read(file_path: str, metadata: dict) -> Tuple[str, None, dict]:
    """Fallback: try to read as plain text. Only safe for text-based formats."""
    path = Path(file_path)
    metadata["conversion"] = "fallback_text"

    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            content = path.read_text(encoding=encoding)
            # Strip NUL bytes defensively
            content = content.replace("\x00", "")
            metadata["encoding"] = encoding
            return content, None, metadata
        except (UnicodeDecodeError, Exception):
            continue

    return f"[Could not read file: {path.name}]", None, metadata
